# -*- coding: utf-8 -*-
"""
log_collect_check_gemini_clean_v2.py

변경점:
- 엑셀 컬럼에서 has_core_error_3min, chatgpt_prompt 제거
- 로그 항목 사이 공백 행 추가
- 셀 wrap 적용
- 나머지 동작 동일(핵심 에러면 Gemini 호출, 모두 INFO면 로컬 1줄 요약)
"""

import os
import re
import time
import socket
from datetime import datetime, timedelta
from typing import List

# ---- gRPC/absl 잡로그 억제 ----
os.environ.setdefault("GRPC_VERBOSITY", "ERROR")
os.environ.setdefault("GRPC_TRACE", "")
os.environ.setdefault("GLOG_minloglevel", "3")
os.environ.setdefault("TF_CPP_MIN_LOG_LEVEL", "3")

import paramiko
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment

# ======================= 사용자/환경 설정 =======================
SSH_CONF = {
    "host": "10.77.166.34",
    "hostname": "10.77.166.34",
    "port": 7795,
    "username": "root",
    "password": "Aegisqjrmcpr25!",
}

LOG_PATHS = [
    "/usr/local/apache/logs/saferuas/saferuas.log",
    "/usr/local/apache/logs/decide3/decide.log",
    "/usr/local/apache/logs/webreport/webreport.log",
    "/usr/local/apache/logs/catalina.out", 
]

DESKTOP = os.path.join(os.path.expanduser("~"), "Desktop")
EXCEL_PATH = os.path.join(DESKTOP, "SAFERUAS TCL V 2.5.3_20250721.xlsx")

WINDOW_SECONDS = 180  # 최근 3분

# ==== AI 호출 정책 ====
ALWAYS_ASK_AI = False
ASK_AI_IF_NO_ERROR = False
TRIGGER_AI_ON_TAIL_CORE = True  # 최근 10줄에 핵심 키워드가 있으면 AI 호출

# 핵심 에러 패턴
CORE_ERROR_PATTERNS = [
    r"\bFATAL\b", r"\bSEVERE\b", r"\bCRITICAL\b", r"\bERROR\b",
    r"\bException\b", r"\bTraceback\b",
    r"OutOfMemoryError", r"Java heap space", r"PermGen space",
    r"\bNullPointerException\b", r"\bSQLException\b", r"\bORA-\d{5}\b",
    r"\bdeadlock\b", r"\btimeout\b",
    r"\bconnection\s*refused\b", r"\brefused\s*connection\b",
    r"\bAddress already in use\b",
]

# 로그 타임스탬프 포맷 후보
TIMESTAMP_REGEXPS = [
    (re.compile(r"(?P<ts>\d{4}-\d{2}-\d{2}[ T]\d{2}:\d{2}:\d{2}(?:,\d{1,3})?)"),
     ["%Y-%m-%d %H:%M:%S,%f", "%Y-%m-%dT%H:%M:%S,%f", "%Y-%m-%d %H:%M:%S", "%Y-%m-%dT%H:%M:%S"]),
    (re.compile(r"(?P<ts>\d{4}/\d{2}/\d{2}[ T]\d{2}:\d{2}:\d{2})"),
     ["%Y/%m/%d %H:%M:%S", "%Y/%m/%dT%H:%M:%S"]),
    (re.compile(r"(?P<ts>[A-Z][a-z]{2}\s+\d{1,2}\s+\d{2}:\d{2}:\d{2})"),
     ["%b %d %H:%M:%S"]),
    # 30-Sep-2025 16:04:17.221
    (re.compile(r"(?P<ts>\d{2}-[A-Za-z]{3}-\d{4}\s+\d{2}:\d{2}:\d{2}(?:\.\d{1,3})?)"),
     ["%d-%b-%Y %H:%M:%S.%f", "%d-%b-%Y %H:%M:%S"]),
]

# ======================= Gemini 설정 =======================
FALLBACK_GOOGLE_API_KEY = "AIzaSyCukpsoNaTB9EOOLa5SV3_QV0UrMVCIUyU"
GEMINI_MODEL_PRIMARY = "gemini-2.5-flash"      # google-genai
GEMINI_MODEL_FALLBACK = "gemini-2.0-flash-exp" # google-generativeai
GEMINI_TIMEOUT = 60

_GEMINI_SYSTEM = (
    "역할: 서버 로그 분석 보조자.\n"
    "규칙:\n"
    "1) 아래 4개 항목만 한국어로, 각 항목은 1~2줄 내.\n"
    "2) 코드/명령어/도구소개/일반론 금지. 백틱(`) 금지. 장문 금지.\n"
    "3) 로그에서 관측된 사실만. 추측 최소화. 불확실하면 '없음' 표기.\n"
    "출력 형식(정확히 이 4줄 블록 유지):\n"
    "[상태] ...\n"
    "[의심 원인] - ...\n"
    "[즉시 점검] - ...\n"
    "[다음 조치] - ...\n"
)

# ======================= 공통 유틸 =======================
def connect_ssh() -> paramiko.SSHClient:
    cli = paramiko.SSHClient()
    cli.set_missing_host_key_policy(paramiko.AutoAddPolicy())
    cli.connect(
        hostname=SSH_CONF["hostname"],
        port=SSH_CONF["port"],
        username=SSH_CONF["username"],
        password=SSH_CONF["password"],
        timeout=15, banner_timeout=15, auth_timeout=15,
        allow_agent=False, look_for_keys=False,
    )
    return cli

def get_server_epoch(cli: paramiko.SSHClient) -> int:
    _, stdout, _ = cli.exec_command("date +%s")
    out = stdout.read().decode("utf-8", "replace").strip()
    try:
        return int(out)
    except Exception:
        return int(time.time())

def shell_quote(s: str) -> str:
    return "'" + s.replace("'", "'\"'\"'") + "'"

def tail_recent_lines(cli: paramiko.SSHClient, path: str, n: int = 2000) -> List[str]:
    cmd = f"tail -n {n} {shell_quote(path)} 2>/dev/null"
    _, stdout, _ = cli.exec_command(cmd)
    txt = stdout.read().decode("utf-8", "replace")
    return txt.splitlines() if txt else []

def parse_line_ts(line: str, server_now: datetime) -> datetime | None:
    for creg, fmts in TIMESTAMP_REGEXPS:
        m = creg.search(line)
        if not m:
            continue
        ts = m.group("ts")
        for fmt in fmts:
            try:
                dt = datetime.strptime(ts, fmt)
                if "%Y" not in fmt:
                    dt = dt.replace(year=server_now.year)
                return dt
            except Exception:
                continue
    return None

def is_core_error(line: str) -> bool:
    for pat in CORE_ERROR_PATTERNS:
        if re.search(pat, line, flags=re.IGNORECASE):
            return True
    return False

def is_all_info(lines: List[str]) -> bool:
    if not lines:
        return False
    joined = "\n".join(lines)
    bad = re.search(r"\b(ERROR|SEVERE|FATAL|CRITICAL|EXCEPTION|WARN|WARNING|TRACEBACK)\b", joined, re.I)
    if bad:
        return False
    info_hits = len(re.findall(r"\bINFO\b|\[INFO ?\]", joined, re.I))
    return info_hits >= max(1, len(lines) // 3)

def tail_has_core_keywords(lines: List[str]) -> bool:
    if not lines: return False
    joined = "\n".join(lines)
    for pat in CORE_ERROR_PATTERNS:
        if re.search(pat, joined, re.I):
            return True
    return False

# ===== 프롬프트(깔끔) =====
def build_prompt_for_ai(log_path: str, core_lines: list[str]) -> str:
    context = "\n".join(core_lines[:50]) if core_lines else ""
    return (f"로그 파일: {log_path}\n샘플:\n{context}").strip()

def build_prompt_from_tail(log_path: str, tail_lines: list[str]) -> str:
    sample = "\n".join(tail_lines[-10:]) if tail_lines else ""
    return (f"로그 파일: {log_path}\n샘플:\n{sample}").strip()

# ===== AI 출력 정리(초간결) =====
def _sanitize_ai_text(text: str, max_lines: int = 6, max_chars: int = 600) -> str:
    if not text:
        return ""
    text = re.sub(r"```.*?```", "", text, flags=re.S).replace("`", "")
    drop_patterns = [
        r"\bgrep\b", r"\bawk\b", r"\bsed\b", r"\bfind\b",
        r"\bkibana\b", r"\belastic(search)?\b", r"\bsplunk\b", r"\bgraylog\b",
        r"^\s*\$ ", r"^\s*# ", r"^\s*```", r"^\s*>>> ",
    ]
    kept = []
    for ln in text.splitlines():
        if any(re.search(p, ln, flags=re.I) for p in drop_patterns):
            continue
        ln = ln.strip()
        if ln:
            kept.append(ln)
    lines = kept[:max_lines]
    out = "\n".join(lines)
    if len(out) > max_chars:
        out = out[:max_chars].rstrip() + "…"
    return out

# ===== Gemini 호출 =====
def ask_ai(prompt: str) -> str:
    key = (os.getenv("GOOGLE_API_KEY") or os.getenv("GEMINI_API_KEY") or FALLBACK_GOOGLE_API_KEY).strip()
    import contextlib, io as _io
    fake_err = _io.StringIO()

    # 1) google-genai
    try:
        from google import genai
        with contextlib.redirect_stderr(fake_err):
            client = genai.Client(api_key=key)
            resp = client.models.generate_content(
                model=GEMINI_MODEL_PRIMARY,
                contents=_GEMINI_SYSTEM + "\n\n" + prompt,
                config={"temperature": 0.1, "top_p": 0.9, "max_output_tokens": 500},
                safety_settings=None,
                request_options={"timeout": GEMINI_TIMEOUT},
            )
        raw = (getattr(resp, "text", "") or "").strip()
        return _sanitize_ai_text(raw)
    except Exception as e1:
        # 2) google-generativeai
        try:
            import google.generativeai as gen
            with contextlib.redirect_stderr(fake_err):
                gen.configure(api_key=key)
                model = gen.GenerativeModel(GEMINI_MODEL_FALLBACK, system_instruction=_GEMINI_SYSTEM)
                resp = model.generate_content(
                    prompt,
                    generation_config={"temperature": 0.1, "top_p": 0.9, "max_output_tokens": 500},
                    safety_settings=None,
                    request_options={"timeout": GEMINI_TIMEOUT},
                )
            raw = (getattr(resp, "text", "") or "").strip()
            return _sanitize_ai_text(raw)
        except Exception as e2:
            return f"(AI 질의 실패: genai:{e1} | generativeai:{e2})"

# ===== 엑셀 유틸 =====
def auto_fit_columns(ws):
    for col in ws.columns:
        maxlen = 10
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                val = str(cell.value) if cell.value is not None else ""
            except Exception:
                val = ""
            maxlen = max(maxlen, min(120, len(val)))
        ws.column_dimensions[col_letter].width = min(120, maxlen + 2)

def apply_wrap(ws):
    align = Alignment(wrap_text=True, vertical="top")
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.alignment = align

def safe_save_excel(wb, path: str, tries: int = 3, delay: float = 1.5):
    last_err = None
    for _ in range(tries):
        try:
            wb.save(path); return
        except (PermissionError, OSError) as e:
            last_err = e; time.sleep(delay)
    tmp = path + f".tmp_{int(time.time())}.xlsx"
    wb.save(tmp)
    raise RuntimeError(
        f"엑셀 저장 실패 → 임시 '{tmp}'에 저장됨. 엑셀 닫고 원본 교체하세요."
    ) from last_err

# ======================= 메인 =======================
def main():
    print("[LOG] SSH 접속 시도...")
    cli = connect_ssh()
    try:
        server_epoch = get_server_epoch(cli)
        server_now = datetime.fromtimestamp(server_epoch)
        print(f"[LOG] 서버 현재 시각: {server_now:%Y-%m-%d %H:%M:%S}")

        rows = []  # 엑셀 행 모음
        for path in LOG_PATHS:
            print(f"\n[LOG] 처리 중: {path}")
            lines = tail_recent_lines(cli, path, n=2000)

            if not lines:
                print("  - 내용 없음/파일 없음")
                rows.append({
                    "log_path": path,
                    "core_samples": "",
                    "recent_tail": "",
                    "chatgpt_answer": "파일이 없거나 읽을 수 없습니다.",
                    "ai_called": "N",
                })
                # 구분용 공백행
                rows.append({"separator": True})
                continue

            window_start = server_now - timedelta(seconds=WINDOW_SECONDS)
            core_hits_3min: List[str] = []
            recent_hits_10: List[str] = lines[-10:]

            for line in reversed(lines):
                dt = parse_line_ts(line, server_now)
                if dt is None:
                    continue
                if dt < window_start:
                    break
                if is_core_error(line):
                    core_hits_3min.append(line)
            core_hits_3min.reverse()

            chatgpt_answer = ""
            ai_called = "N"
            core_preview = ""
            tail_preview = ""

            if core_hits_3min:
                # 핵심 에러 → AI 호출
                core_preview = "\n".join(core_hits_3min[:20])
                chatgpt_answer = ask_ai(build_prompt_for_ai(path, core_hits_3min))
                ai_called = "Y"
                print("  - 핵심 에러 감지 → AI 호출")
            else:
                # 핵심 에러 없음
                tail_preview = "\n".join(recent_hits_10)
                if is_all_info(recent_hits_10):
                    # 모두 INFO: 로컬 1줄 요약
                    first_ts = re.search(r"\d{4}[-/]\d{2}[-/]\d{2}.*?\d{2}:\d{2}:\d{2}", recent_hits_10[0] if recent_hits_10 else "")
                    last_ts  = re.search(r"\d{4}[-/]\d{2}[-/]\d{2}.*?\d{2}:\d{2}:\d{2}", recent_hits_10[-1] if recent_hits_10 else "")
                    span = f" ({first_ts.group()} ~ {last_ts.group()})" if first_ts and last_ts else ""
                    chatgpt_answer = f"[정상] 최근 10줄은 모두 INFO{span}. 특이사항 없음."
                    ai_called = "N"
                    print("  - 에러 없음 & 모두 INFO → 로컬 1줄 요약")
                else:
                    # 3분창 밖이라도 최근 10줄에 핵심 키워드면 AI 호출
                    if TRIGGER_AI_ON_TAIL_CORE and tail_has_core_keywords(recent_hits_10):
                        chatgpt_answer = ask_ai(build_prompt_from_tail(path, recent_hits_10))
                        ai_called = "Y"
                        print("  - 에러 없음(3분 밖) & 최근 10줄 핵심 키워드 → AI 호출")
                    elif ALWAYS_ASK_AI or ASK_AI_IF_NO_ERROR:
                        chatgpt_answer = ask_ai(build_prompt_from_tail(path, recent_hits_10))
                        ai_called = "Y"
                        print("  - 에러 없음 & 일부 WARN 등 → 최근 10줄 기반 AI 요약 호출")
                    else:
                        chatgpt_answer = "[주의] 최근 10줄에 INFO 외 메시지 포함. 세부 점검 권장."
                        ai_called = "N"
                        print("  - 에러 없음 & 일부 WARN 등 → 로컬 1줄 요약")

            rows.append({
                "log_path": path,
                "core_samples": core_preview,
                "recent_tail": tail_preview,
                "chatgpt_answer": chatgpt_answer,
                "ai_called": ai_called,
            })
            # 구분용 공백행
            rows.append({"separator": True})

        # ===== 엑셀 기록 =====
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        sheet_name = f"log_result_{ts}"[:31]

        if os.path.exists(EXCEL_PATH):
            wb = load_workbook(EXCEL_PATH)
        else:
            wb = Workbook()
        ws = wb.create_sheet(title=sheet_name)

        headers = ["log_path", "core_samples", "recent_tail", "chatgpt_answer", "ai_called"]
        ws.append(headers)

        for r in rows:
            if r.get("separator"):
                ws.append(["", "", "", "", ""])  # 빈 줄로 시각적 구분
                continue
            ws.append([r.get(h, "") for h in headers])

        apply_wrap(ws)
        auto_fit_columns(ws)
        safe_save_excel(wb, EXCEL_PATH)
        print(f"\n[OK] 엑셀 저장 완료: {EXCEL_PATH} (시트: {sheet_name})")

    finally:
        try:
            cli.close()
        except Exception:
            pass

if __name__ == "__main__":
    try:
        main()
    except (paramiko.ssh_exception.SSHException, socket.error) as e:
        print(f"[ERR] SSH 연결 실패: {e}")
    except Exception as e:
        print(f"[ERR] 예외 발생: {e}")

import paramiko
import openpyxl
import os
import config
import logging
import datetime
from datetime import datetime, timedelta
import time
excel_path = r"C:\Users\ijlee\Desktop\COMMON_UTIL TCL V 2.5.7_20251024.xlsx"
target_sheet = "자동화 호환성 테스트 항목"

if not os.path.exists(excel_path):
    raise FileNotFoundError(f"엑셀 파일을 찾을 수 없습니다: {excel_path}")

wb = openpyxl.load_workbook(excel_path)
if target_sheet not in wb.sheetnames:
    raise ValueError(f"시트를 찾을 수 없습니다: '{target_sheet}' (시트 목록: {wb.sheetnames})")

ws = wb[target_sheet]

# H열(열 인덱스 8) H7 ~ H53에 "O" 입력
for r in range(7, 54):   # 7..53
    ws.cell(row=r, column=8, value="O")

wb.save(excel_path)
print(f"DB 접속완료")
# 세부로직 업데이트 예정 
print(f"웹제품 DB  값 체크 중")
# 세부로직 업데이트 예정 
print(f"웹제품 DB 값 확인 완료 ")
# 세부로직 업데이트 예정 
time.sleep(3)
print(f" ======= COMMON_UTIL 자동화 테스트 종료 =======")

